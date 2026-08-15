import csv, io, re
from datetime import datetime
from fastapi import HTTPException
from sqlalchemy.orm import Session
from .auth import normalize_phone
from ..models import (
    User, Course, Enrollment, StudentProfile, StudentGroup, StudentGroupMembership,
    GroupCourseAssignment, ActiveSession, CommunicationCampaign, CommunicationDelivery,
    Notification,
)
from ..security import hash_password


def set_student_group(db: Session, student_id: int, group_id: int | None):
    row = db.query(StudentGroupMembership).filter_by(user_id=student_id).first()
    if not group_id:
        if row:
            db.delete(row)
        return
    group = db.get(StudentGroup, group_id)
    if not group:
        raise HTTPException(404, "المجموعة غير موجودة")
    if row:
        row.group_id = group_id
        row.joined_at = datetime.utcnow()
    else:
        db.add(StudentGroupMembership(group_id=group_id, user_id=student_id))
    db.flush()
    for assignment in db.query(GroupCourseAssignment).filter_by(group_id=group_id).all():
        enrollment = db.query(Enrollment).filter_by(user_id=student_id, course_id=assignment.course_id).first()
        if enrollment:
            enrollment.active = True
        else:
            db.add(Enrollment(user_id=student_id, course_id=assignment.course_id, active=True))


def import_student_rows(raw: bytes, filename: str):
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(500, "دعم Excel غير مثبت على الخادم") from exc
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(x or "").strip() for x in values[0]]
        return [dict(zip(headers, row)) for row in values[1:] if any(v not in (None, "") for v in row)]
    raise HTTPException(400, "ارفع ملف CSV أو XLSX فقط")


def row_value(row, *names):
    normalized = {str(k).strip().lower(): v for k, v in row.items()}
    for name in names:
        value = normalized.get(name.lower())
        if value not in (None, ""):
            return str(value).strip()
    return ""


def validate_admin_password(password: str, *, default: bool = False):
    if len(password) < 12 or not re.search(r"[A-Za-z]", password) or not re.search(r"\d", password):
        if default:
            raise HTTPException(400, "كلمة المرور الافتراضية يجب أن تكون 12 حرفًا على الأقل وتحتوي حروفًا وأرقامًا")
        raise HTTPException(400, "كلمة المرور يجب أن تكون 12 حرفًا على الأقل وتحتوي حروفًا وأرقامًا")


def revoke_user_sessions(db: Session, user_id: int):
    db.query(ActiveSession).filter(
        ActiveSession.user_id == user_id,
        ActiveSession.revoked_at.is_(None),
    ).update({ActiveSession.revoked_at: datetime.utcnow()}, synchronize_session=False)
